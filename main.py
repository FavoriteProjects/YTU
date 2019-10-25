import openpyxl

import os

import google.oauth2.credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.http import MediaFileUpload


# wb_name="Video_List.xlsx"
wb_name = "list.xlsx"
video_wb = openpyxl.load_workbook(wb_name)
sheet = video_wb.active


def Upload_video(service, title, description, video_source):
    request = service.videos().insert(
        part="snippet,status",
        autoLevels=True,
        notifySubscribers=True,
        stabilize=True,
        body={
            "snippet": {
                # "categoryId": "22",
                "description": abstract,
                "title": title
            },
            "status": {
                "privacyStatus": "public"
            }
        },

        media_body=MediaFileUpload(
            video_source, mimetype='video/mp4', resumable=True)
    )
    response = None
    while response is None:
        status, response = request.next_chunk()
        if status:
            print(("Uploaded %d%%.") % int(status.progress() * 100))
    print("Upload Complete!")
    response = status

    return "done"


CLIENT_SECRETS_FILE = "client_secrets.json"
SCOPES = ["https://www.googleapis.com/auth/youtube.upload"]
API_SERVICE_NAME = "youtube"
API_VERSION = 'v3'


def get_authenticated_service():
    flow = InstalledAppFlow.from_client_secrets_file(
        CLIENT_SECRETS_FILE, SCOPES)
    credentials = flow.run_console()
    return build(API_SERVICE_NAME, API_VERSION, credentials=credentials)


if __name__ == '__main__':
      # When running locally, disable OAuthlib's HTTPs verification. When
      # running in production *do not* leave this option enabled.
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    service = get_authenticated_service()
    print(service)

    maxR = sheet.max_row

    for i in range(2, maxR+1):
        if(sheet.cell(row=i, column=8).value == 0):
            title = sheet.cell(row=i, column=2).value
            abstract = sheet.cell(row=i, column=6).value
            author = sheet.cell(row=i, column=3).value
            affiliation = sheet.cell(row=i, column=4).value
            session = sheet.cell(row=i, column=5).value
            description = "abstract: "+abstract+"\n"+"author: "+author + \
                "\n"+"affiliation: "+affiliation+"\n"+"session: "+session
            video_source = str(sheet.cell(row=i, column=1).value)+".mp4"
            status = Upload_video(service, title, description, video_source)
            print("Uploaded"+" "+str(i-1))
            if(status == 'done'):
                sheet.cell(row=i, column=8).value = 1
                video_wb.save(wb_name)
